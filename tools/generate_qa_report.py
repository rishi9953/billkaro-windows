"""Generate Billkaro QA Excel report (post-fix update)."""
from __future__ import annotations

from datetime import datetime
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill

HEADERS = [
    "ID",
    "Category",
    "Severity",
    "Module / Feature",
    "Title",
    "Description",
    "Steps to Reproduce",
    "Expected Behavior",
    "Actual Behavior (Before Fix)",
    "Suggested Fix",
    "File Path(s)",
    "Status",
    "Resolution / Fix Applied",
    "Priority Rank",
]

SEVERITY_COLORS = {
    "Critical": "FFC7CE",
    "High": "FFE699",
    "Medium": "FFF2CC",
    "Low": "E2EFDA",
}

STATUS_COLORS = {
    "Fixed": "C6EFCE",
    "Partial": "FFF2CC",
    "Open": "FFC7CE",
    "Deferred": "E2EFDA",
}

# id -> (status, resolution)
RESOLUTIONS: dict[str, tuple[str, str]] = {
    "QA-001": ("Fixed", "login_controller: showError on null response and in catch block."),
    "QA-002": ("Fixed", "Password minimum raised to 8 characters in validateDeviceLabel()."),
    "QA-003": ("Fixed", "Validation messages updated to email/password field labels."),
    "QA-004": ("Fixed", "Login blocked with error when outletData is empty."),
    "QA-005": ("Fixed", "Login screen link text changed to 'Forgot password?'."),
    "QA-006": ("Fixed", "onAddDevice() no longer fakes success; shows not-available message."),
    "QA-007": ("Fixed", "SSL bypass limited to kDebugMode only (main.dart + api_overRides.dart)."),
    "QA-008": ("Fixed", "StaffAccess helper, sidebar route guards, staff permissions in login + backend."),
    "QA-009": ("Open", "Not addressed — splash still uses fixed delay."),
    "QA-010": ("Fixed", "Offline save catch now logs error and shows message with exception detail."),
    "QA-011": ("Fixed", "SyncManager.triggerSync(immediate: true) after offline save when online."),
    "QA-012": ("Open", "Not addressed — offline multi-terminal table locking still needs work."),
    "QA-013": ("Fixed", "kot_preview_controller getUserDetails() made null-safe with fallbacks."),
    "QA-014": ("Fixed", "KotPrintJob.buildPdf() implemented with basic KOT PDF layout."),
    "QA-015": ("Open", "Order preferences navigation still commented out."),
    "QA-016": ("Fixed", "Item reports payment filter uses ?.toLowerCase() null-safe check."),
    "QA-017": ("Open", "Customer filter in reports still not implemented."),
    "QA-018": ("Fixed", "Business overview uses Modular.to.navigate(HomeMainRoutes.itemsReport)."),
    "QA-019": ("Open", "Order reports PDF preview route still commented."),
    "QA-020": ("Fixed", "Order reports guards userId/outletId before API call."),
    "QA-021": ("Fixed", "Removed fake _defaultTables(); shows empty/error state instead."),
    "QA-022": ("Fixed", "closeOrderAndFreeTable() updates local DB via updateOrderStatus()."),
    "QA-023": ("Fixed", "KDS WebSocket defaults to REST host; local WS opt-in via .env only."),
    "QA-024": ("Fixed", "KDS disconnect errors logged instead of empty catch blocks."),
    "QA-025": ("Fixed", "KOT history loadError shown; loadMore shows snackbar on failure."),
    "QA-026": ("Open", "Route typo whatsaapMarketing unchanged (breaking change if renamed)."),
    "QA-027": ("Fixed", "Windows printer skips BT error; defaults to USB tab on launch."),
    "QA-028": ("Open", "Some printer empty catch blocks remain in thermal_printer_service."),
    "QA-029": ("Fixed", "KOT validate() no longer requires waiterName; PDF uses Staff fallback."),
    "QA-030": ("Fixed", "Customer list: loading, error/retry, and empty states separated."),
    "QA-031": ("Fixed", "Customer API uses response?.status; no force-unwrap crash."),
    "QA-032": ("Open", "CutomerListController typo not renamed (refactor deferred)."),
    "QA-033": ("Fixed", "Edit menu maps to edit_menu permission; backend DTO updated."),
    "QA-034": ("Fixed", "Staff edit hydration reads edit_menu permission for checkbox."),
    "QA-035": ("Open", "Signup address save still local-only."),
    "QA-036": ("Fixed", "WhatsApp uses Twilio + outlet customer list; ngrok/hardcoded numbers removed."),
    "QA-037": ("Fixed", "sendBulkWhatsAppMessages() restored with Twilio integration."),
    "QA-038": ("Open", "Language menu entry still commented out."),
    "QA-039": ("Open", "Branding strings not standardized."),
    "QA-040": ("Open", "OrderPrefrences folder typo not renamed."),
    "QA-041": ("Fixed", "syncPendingOrders syncs all pending orders, not selected outlet only."),
    "QA-042": ("Fixed", "Windows sync uses in-app showSuccess/showError snackbars."),
    "QA-043": ("Partial", "4xx orders marked failed in DB + user notified; retry queue UI not built."),
    "QA-044": ("Fixed", "502 no longer calls _handleTokenExpiration(); shows retry message."),
    "QA-045": ("Fixed", "Default API URL set to production; ApiConstants.defaultBase = prod."),
    "QA-046": ("Fixed", "billkaro_windows widget test replaced; 1/1 smoke test passes."),
    "QA-047": ("Open", "Auth flow unit tests not added beyond smoke test."),
    "QA-048": ("Open", "Offline sync integration tests not added."),
    "QA-049": ("Open", "Reports export/filter tests not added."),
    "QA-050": ("Open", "Printer/KDS unit tests not added."),
    "QA-051": ("Fixed", "NetworkUtils uses result.contains(ConnectivityResult.none)."),
    "QA-052": ("Fixed", "app_snackbar.dart copied to billkaro-app; compile errors resolved."),
    "QA-053": ("Open", "Static analysis warnings remain (~1400 info items)."),
    "QA-054": ("Open", "Dependency upgrades not performed."),
    "QA-055": ("Open", "CMake/NuGet build warnings remain."),
    "QA-056": ("Open", "home_screen.dart not refactored."),
    "QA-057": ("Open", "add_order_controller.dart not split into services."),
    "QA-058": ("Partial", "Keyboard workaround retained in main.dart; upstream Flutter issue."),
    "QA-059": ("Open", "Windows background sync via Workmanager still unavailable."),
    "QA-060": ("Open", "Full E2E customer journey tests not added."),
}


def f(*cols):
    return list(cols)


FINDINGS = [
    f("QA-001", "Bug", "Critical", "Login", "Silent login failure",
      "Login only navigates when API response is non-null. Catch block only logs to console.",
      "1. Enter credentials\n2. API returns null or throws non-Dio error",
      "Clear error snackbar/dialog shown", "Spinner stops; user stays on login with no message",
      "Show showError() when response is null and in catch block",
      r"lib\app\modules\Login\login_controller.dart", "2"),
    f("QA-002", "Bug", "High", "Login", "Weak password validation",
      "Password validator allows minimum 3 characters.", "Sign in with 3-character password",
      "Stronger password policy (8+ chars)", "Accepts 3-character passwords",
      "Align validation with backend policy", r"lib\app\modules\Login\login_controller.dart", "15"),
    f("QA-003", "Bug", "High", "Login", "Misleading validation messages",
      "Validators show email/password errors for registration key / device label fields.",
      "Submit login form with empty fields", "Field-specific error labels", "Confusing copied validation messages",
      "Rename validators and error strings to match login fields", r"lib\app\modules\Login\login_controller.dart", "20"),
    f("QA-004", "Bug", "Medium", "Login", "Login proceeds without outlet",
      "App navigates to home even when outletData is null or empty.",
      "Log in with account that has no outlets configured", "Outlet setup prompt or error",
      "Lands in app with no selected outlet", "Block navigation and show setup flow when no outlets",
      r"lib\app\modules\Login\login_controller.dart", "18"),
    f("QA-005", "Bug", "Medium", "Login", "Label mismatch: Request registration key",
      "Button says Request registration key but opens Forgot Password flow.",
      "Tap link on login screen", "Label matches behavior", "Mismatched label and action",
      "Rename to Forgot password?", r"lib\app\modules\Login\login_screen.dart", "25"),
    f("QA-006", "Bug", "Medium", "Login", "Stub Add Device shows fake success",
      "onAddDevice() simulates delay and shows success without API call.",
      "Submit add-device form if exposed", "Real device registration via API", "Fake success after 2s delay",
      "Wire to real API or remove/hide feature", r"lib\app\modules\Login\login_controller.dart", "30"),
    f("QA-007", "Bug", "Critical", "Security", "Global SSL certificate bypass",
      "MyHttpOverrides accepts all TLS certificates; set globally in main and login.",
      "MITM on public/corporate Wi-Fi", "Valid certificate validation", "All certificates trusted",
      "Remove bypass in production", r"lib\app\services\Network\api_overRides.dart; lib\main.dart", "1"),
    f("QA-008", "Bug", "High", "Auth / Staff", "Staff session flag never enforced",
      "isStaffSession is set on login but never checked elsewhere for RBAC.",
      "Log in as staff; access admin screens", "Role-based route/action restrictions", "No RBAC enforcement in UI",
      "Gate routes on isStaffSession and permissions", r"lib\utils\staff_access.dart; app_shell_sidebar.dart", "8"),
    f("QA-009", "Improvement", "Low", "Splash", "Fixed 3-second splash delay",
      "Splash always waits 3 seconds before routing regardless of init speed.",
      "Launch app", "Navigate when initialization completes", "Artificial 3s wait every launch",
      "Route as soon as auth/prefs are ready", r"lib\app\modules\splash\splash_controller.dart", "45"),
    f("QA-010", "Bug", "High", "Add Order", "Generic offline save failure message",
      "Offline branch catch shows generic failed_to_save_order_offline without root cause.",
      "Save order offline when DB write fails", "Actionable error with reason", "Generic offline failure message only",
      "Log and surface specific failure reason", r"lib\app\modules\AddOrder\add_order_controller.dart", "14"),
    f("QA-011", "Improvement", "Medium", "Add Order / Sync", "Offline orders do not trigger immediate sync",
      "Offline save writes to DB but does not call SyncManager.triggerSync() when back online from this flow.",
      "Create offline order; reconnect network", "Prompt sync attempt on reconnect",
      "Relies on 15-min timer or connectivity listener only",
      "Trigger sync after successful offline save when online",
      r"lib\app\modules\AddOrder\add_order_controller.dart; sync_manager.dart", "22"),
    f("QA-012", "Bug", "Medium", "Add Order / Tables", "Dine-in table conflict edge cases offline",
      "Active table order check may desync across offline terminals.",
      "Two terminals offline assign same table", "Consistent conflict prevention", "Depends on local state freshness",
      "Strengthen offline table locking and conflict UI", r"lib\app\modules\AddOrder\add_order_controller.dart", "28"),
    f("QA-013", "Bug", "Medium", "KOT Preview", "Null-force crash on staff/user",
      "getUserDetails() uses appPref.user! and outletData!.first without null checks.",
      "Open KOT preview with incomplete profile", "Graceful fallback names", "Possible runtime crash",
      "Null-safe access with fallback values", r"lib\app\modules\Invoice\KOT\kot_preview_controller.dart", "16"),
    f("QA-014", "Bug", "Medium", "KOT Print", "KOT PDF export unimplemented",
      "buildPdf() throws UnimplementedError.", "Trigger KOT PDF export path", "Working PDF or disabled UI",
      "Runtime UnimplementedError", "Implement PDF generation",
      r"lib\app\services\printerService.dart\thermal_printer\Print\kot_print.dart", "24"),
    f("QA-015", "Improvement", "Low", "Add Order", "Order preferences navigation commented out",
      "Order preferences route call is commented out in controller.",
      "Try accessing preferences from add-order flow", "Working preferences shortcut", "Dead code path",
      "Restore navigation or remove UI entry", r"lib\app\modules\AddOrder\add_order_controller.dart", "50"),
    f("QA-016", "Bug", "High", "Item Reports", "Null crash on payment filter",
      "Offline filter uses order.paymentReceivedIn!.toLowerCase() — crashes when null.",
      "Filter Item Reports by payment type with null payment field", "Null-safe filter match",
      "Null assertion crash", "Use ?.toLowerCase() null-safe check",
      r"lib\app\modules\Reports\ItemReports\item_reports_controller.dart", "7"),
    f("QA-017", "Improvement", "Medium", "Reports", "Customer filter not implemented",
      "filterByCustomers() shows coming soon; selectedCustomers filter state unused in UI.",
      "Open reports and look for customer filter", "Working customer filter dialog",
      "Stub message; dead backend filter state", "Implement dialog or remove filter UI/state",
      r"lib\app\modules\Reports\ItemReports\item_reports_controller.dart", "32"),
    f("QA-018", "Bug", "Medium", "Business Overview", "Wrong navigation to Item Reports",
      "Uses Get.toNamed(AppRoute.itemReports) instead of Modular shell route.",
      "From Business Overview tap View Item Reports", "Opens item reports inside shell",
      "May push outside Modular context", "Use Modular.to.navigate(HomeMainRoutes.itemsReport)",
      r"lib\app\modules\BusinessOverview\business_overview_screen.dart", "26"),
    f("QA-019", "Improvement", "Low", "Order Reports", "PDF preview route commented out",
      "Large block of invoice PDF preview navigation is commented.",
      "Tap invoice action in order reports if visible", "PDF preview opens", "Incomplete path only",
      "Finish PDF preview or remove dead code", r"lib\app\modules\Reports\OrderReports\order_reports_screen.dart", "48"),
    f("QA-020", "Bug", "Medium", "Order Reports", "Hard null on user ID fetch",
      "API call uses appPref.user!.id! without guard.", "Open reports with cleared user pref",
      "Error message shown", "Possible crash", "Guard user/outlet before API calls",
      r"lib\app\modules\Reports\OrderReports\order_reports_controller.dart", "21"),
    f("QA-021", "Bug", "High", "Tables", "Fake default tables on API failure",
      "When API fails, _defaultTables() generates 12 local phantom tables.",
      "Open Tables with API down", "Empty state or error message",
      "12 invented tables shown", "Show error/empty state; do not invent tables",
      r"lib\app\modules\Tables\table_controller.dart", "6"),
    f("QA-022", "Bug", "High", "Tables", "Table free-after-payment incomplete",
      "closeOrderAndFreeTable() DB update was commented out.",
      "Complete payment and free table", "Order closed in DB and table available",
      "Table may remain occupied", "Restore DB update logic",
      r"lib\app\modules\Tables\table_controller.dart", "9"),
    f("QA-023", "Bug", "High", "Kitchen Display", "Windows defaults to localhost WebSocket",
      "KDS WS defaulted to ws://127.0.0.1:3000 while REST used remote URL.",
      "Run KDS on Windows without .env", "WebSocket host matches REST API",
      "WS connects to localhost", "Default local WS to false; opt-in via .env",
      r"lib\app\services\Network\api_config.dart", "11"),
    f("QA-024", "Bug", "Medium", "Kitchen Display", "KDS disconnect errors swallowed",
      "WebSocket close errors caught with empty catch blocks.", "Disconnect during network flap",
      "Surfaced error/reconnect status", "Silent failure", "Log errors; avoid empty catches",
      r"lib\app\services\kds\kds_realtime_service.dart", "33"),
    f("QA-025", "Bug", "Medium", "KOT History", "Load errors hidden from user",
      "load() / loadMore() catch blocks only debugPrint.", "Corrupt DB or load failure",
      "User-visible error with retry", "Silent empty list", "Show error banner and retry",
      r"lib\app\modules\KOTHistory\kot_history_controller.dart", "27"),
    f("QA-026", "Improvement", "Low", "Routes", "Typo in WhatsApp route constant",
      "Route named whatsaapMarketing (misspelled).", "Inspect route names",
      "Consistent whatsappMarketing naming", "Typo persisted in route ID", "Rename with migration plan",
      r"lib\app\modules\HomeMain\home_main_routes.dart", "52"),
    f("QA-027", "Bug", "Medium", "Printer", "Bluetooth not supported error on desktop",
      "checkBluetoothPermission() shows error if BT unsupported.", "Open Printer on Windows",
      "USB-first UX on Windows", "Error snackbar even if USB works", "Gate BT by platform; USB default tab",
      r"lib\app\modules\Printer\printer_controller.dart", "29"),
    f("QA-028", "Bug", "Medium", "Printer", "Empty catch blocks hide print failures",
      "Multiple printer helpers swallow exceptions silently.", "Print or connect failure",
      "User notified of failure", "Silent failure possible", "Surface errors to UI",
      r"lib\app\services\printerService.dart\thermal_printer\thermal_printer_service.dart", "31"),
    f("QA-029", "Bug", "Medium", "KOT Print", "KOT validation requires waiter name",
      "KotPrintJob.validate() fails if waiterName is empty.", "Auto-print KOT when waiter missing",
      "Print with fallback staff name", "Validation failure prevents print", "Relax validation",
      r"lib\app\services\printerService.dart\thermal_printer\Print\kot_print.dart", "34"),
    f("QA-030", "Bug", "High", "Customers", "Empty list shows onboarding not list",
      "customerList.isEmpty renders onboarding for loading/error/zero customers.",
      "Open Customers before API returns or after error", "Loader then list or empty state",
      "Onboarding shown incorrectly", "Add loading/error states",
      r"lib\app\modules\Regular customer\CustomerList\customer_list_Screen.dart", "5"),
    f("QA-031", "Bug", "High", "Customers", "Force-unwrap on failed API",
      "Uses response!.status when callApi returns null.", "Open customers offline",
      "Error message shown", "Null check operator crash", "Use response?.status and error UI",
      r"lib\app\modules\Regular customer\CustomerList\cutomer_list_controller.dart", "4"),
    f("QA-032", "Improvement", "Medium", "Customers", "Typo in controller class name",
      "CutomerListController misspelled throughout module.", "N/A", "CustomerListController naming",
      "Typo in type and file name", "Rename class and file", r"lib\app\modules\Regular customer\CustomerList\cutomer_list_controller.dart", "46"),
    f("QA-033", "Bug", "High", "Staff", "Wrong permission mapping for Edit menu",
      "canEditMenuItems mapped to view_reports not menu edit.", "Create biller with Edit menu checked",
      "Menu edit permission granted", "Only view_reports granted", "Map to edit_menu permission",
      r"lib\app\modules\Staff\add_staff_controller.dart; billkaro-backend outlet-staff DTOs", "10"),
    f("QA-034", "Bug", "Medium", "Staff", "Hydration maps wrong permission to UI",
      "When editing staff, canEditMenuItems set from view_reports.", "Edit existing staff",
      "Checkbox reflects menu edit permissions", "Mislabeled capability", "Fix permission mapping",
      r"lib\app\modules\Staff\add_staff_controller.dart", "23"),
    f("QA-035", "Improvement", "Low", "Signup", "Address save not persisted to backend",
      "AddAddressController.saveAddress() has TODO.", "Complete signup address step",
      "Address saved via API", "Local-only save", "Implement API persistence",
      r"lib\app\modules\Signup\address_controller.dart", "49"),
    f("QA-036", "Bug", "Critical", "WhatsApp Marketing", "Hardcoded dev ngrok URL and phone numbers",
      "Bulk send used ngrok URL with hardcoded test numbers.", "Send WhatsApp campaign",
      "Outlet customer list and Twilio", "Dev tunnel and fixed numbers used",
      "Use Twilio + customer API", r"lib\app\modules\Whatsapp Marketing\whatsapp_marketing_controller.dart", "3"),
    f("QA-037", "Improvement", "Medium", "WhatsApp Marketing", "Primary bulk-send flow commented out",
      "sendBulkWhatsAppMessages fully commented; ngrok workaround active.", "Use Twilio template send",
      "Production WhatsApp integration", "Dead code; ngrok only", "Restore Twilio integration",
      r"lib\app\modules\Whatsapp Marketing\whatsapp_marketing_controller.dart", "35"),
    f("QA-038", "Improvement", "Low", "Settings / Menu", "Language change menu disabled",
      "Language menu item commented out in Menu screen.", "Menu → Language",
      "Language settings accessible", "Entry removed", "Re-enable changeLanguage route",
      r"lib\app\modules\Menu\menu_screen.dart", "47"),
    f("QA-039", "Improvement", "Low", "Branding", "Inconsistent product naming in UI",
      "Mixed Billkaro / BillKaro / ChillKaro spelling in titles.", "View title bar and exit dialog",
      "Consistent brand string", "Mixed spelling", "Standardize branding copy",
      r"lib\main.dart; windows_desktop_title_bar.dart", "51"),
    f("QA-040", "Improvement", "Low", "Module naming", "OrderPrefrences typo in folder/module",
      "Folder uses OrderPrefrences spelling.", "N/A", "OrderPreferences naming",
      "Typo in paths", "Rename module in refactor", "lib\\app\\modules\\OrderPrefrences\\", "53"),
    f("QA-041", "Bug", "High", "Offline Sync", "Only selected outlet pending orders sync",
      "syncPendingOrders filtered to selectedOutletId only.", "Save offline for outlet A; switch to B",
      "All outlets sync or warning", "Other outlet orders stuck pending", "Sync all pending orders",
      r"lib\app\services\Synchronisatioin\synchronisation.dart", "12"),
    f("QA-042", "Bug", "Medium", "Offline Sync", "No Windows sync notifications",
      "SyncNotificationService Android/iOS only.", "Sync on Windows",
      "In-app sync status indicator", "Debug logs only", "Windows snackbar fallback",
      r"lib\app\services\notification\sync_notification_service.dart", "36"),
    f("QA-043", "Bug", "High", "Offline Sync", "4xx failed orders never retried, no user queue",
      "4xx orders logged as will not retry; no queue UI.", "Sync order rejected (4xx)",
      "User notified with fix/retry flow", "Stuck pending silently", "Mark failed + notify user",
      r"lib\app\services\Synchronisatioin\synchronisation.dart; app_database.dart", "13"),
    f("QA-044", "Bug", "High", "Network / Auth", "502 response triggers logout",
      "502 interceptor called _handleTokenExpiration().", "Server returns 502",
      "Retry message; stay logged in", "Session cleared", "Treat 502 as transient",
      r"lib\app\services\Network\network_module.dart", "5"),
    f("QA-045", "Bug", "Critical", "Config", "Default API points to dev tunnel",
      "Fallback base URL was devtunnels.ms when .env missing.", "Run without .env",
      "Production API endpoint", "Dev tunnel used", "Default to production API URL",
      r"lib\app\services\Network\urls.dart; api_config.dart", "1"),
    f("QA-046", "Test Gap", "Critical", "Tests", "Broken default widget test",
      "Test imported wrong package and expected counter app.", "Run flutter test",
      "Passing smoke test", "0/1 compile failure", "Fix test import and expectations",
      r"billkaro_windows\test\widget_test.dart", "1"),
    f("QA-047", "Test Gap", "High", "Tests", "No auth flow tests",
      "No unit tests for login, staff login, token expiry.", "Run flutter test",
      "Login success/failure tests", "Zero coverage", "Add controller tests with mock ApiClient",
      "billkaro_windows\\test\\", "17"),
    f("QA-048", "Test Gap", "High", "Tests", "No order/offline sync tests",
      "No tests for offline save, pending sync, 4xx handling.", "Run flutter test",
      "Sync integration tests", "None exist", "Add in-memory DB + mock API tests",
      "synchronisation.dart; add_order_controller.dart", "19"),
    f("QA-049", "Test Gap", "Medium", "Tests", "No reports export/filter tests",
      "Excel/PDF export and filters untested.", "Run flutter test",
      "Filter/export regression tests", "None exist", "Test applyAllFilters() and exports",
      "item_reports_controller.dart; order_reports_controller.dart", "37"),
    f("QA-050", "Test Gap", "Medium", "Tests", "No printer/KDS integration tests",
      "Printer assignment, KOT validation, KDS reconnect untested.", "Run flutter test",
      "Unit tests for critical paths", "None exist", "Add mocked service tests",
      "printer_controller.dart; kds_realtime_service.dart", "38"),
    f("QA-051", "Bug", "High", "Network", "Connectivity check type mismatch",
      "Compared List<ConnectivityResult> to ConnectivityResult.none.", "Call hasInternetConnection()",
      "Correct none-network detection", "Wrong online/offline state", "Use result.contains(none)",
      r"lib\utils\connectivity\connectivity_helper.dart", "3"),
    f("QA-052", "Test Gap", "High", "billkaro-app", "Missing app_snackbar.dart breaks tests",
      "billkaro-app missing utils/app_snackbar.dart.", "Run flutter test in billkaro-app",
      "All tests compile", "4 compile failures", "Copy app_snackbar.dart from billkaro_windows",
      r"billkaro-app\lib\utils\app_snackbar.dart", "4"),
    f("QA-053", "Improvement", "Medium", "Code Quality", "1441 static analysis issues",
      "Mostly deprecated withOpacity and unnecessary imports.", "Run flutter analyze",
      "Minimal warnings in CI", "~1400 info/warning items", "Enable CI analyze gate incrementally",
      "billkaro_windows\\lib\\", "39"),
    f("QA-054", "Improvement", "Medium", "Dependencies", "107 packages outdated",
      "Many packages behind latest per flutter pub outdated.", "Run flutter pub outdated",
      "Up-to-date dependencies", "Many packages behind", "Plan upgrade sprint",
      r"billkaro_windows\pubspec.yaml", "40"),
    f("QA-055", "Improvement", "Low", "Build", "CMake/NuGet warnings during Windows build",
      "Nuget not installed; webview_windows CMP0175 warning.", "flutter build windows",
      "Clean build without warnings", "Build succeeds with warnings", "Install NuGet; update CMake",
      "build\\windows\\", "54"),
    f("QA-056", "Improvement", "Medium", "Home / Dashboard", "Large home_screen.dart file",
      "Monolithic UI file 1600+ lines.", "Navigate home dashboard",
      "Modular widget composition", "Hard to maintain", "Extract widgets",
      r"lib\app\modules\Home\home_screen.dart", "41"),
    f("QA-057", "Improvement", "Medium", "Add Order", "add_order_controller.dart 2600+ lines",
      "God controller mixing many concerns.", "Create/edit order flow",
      "Separated service layers", "High regression risk", "Split into services",
      r"lib\app\modules\AddOrder\add_order_controller.dart", "42"),
    f("QA-058", "Bug", "Medium", "Keyboard (Windows)", "Hardware keyboard state workaround needed",
      "clearState() workaround for Alt-key stuck after native dialogs.", "Open file dialog then Alt",
      "Normal keyboard state", "Framework assertion without workaround", "Keep workaround; track Flutter fix",
      r"lib\main.dart", "43"),
    f("QA-059", "Improvement", "Low", "Workmanager", "Background sync not available on Windows",
      "Workmanager skipped on Windows.", "Close app with pending offline orders",
      "Background sync when closed", "Sync only when app running", "Document limitation",
      r"lib\main.dart; sync_manager.dart", "44"),
    f("QA-060", "Improvement", "Medium", "Customer UX", "Limited automated E2E tests",
      "No integration test for full customer journey.", "Manual full journey",
      "Automated E2E on CI", "Manual testing only", "Add integration_test package",
      "billkaro_windows\\test\\", "55"),
]


def enrich_rows() -> list[list[str]]:
    rows = []
    for base in FINDINGS:
        qid = base[0]
        status, resolution = RESOLUTIONS.get(qid, ("Open", "Not yet addressed."))
        # base: id..paths (12 cols), priority (last)
        rows.append(base[:11] + [status, resolution, base[11]])
    return rows


def style_header(ws):
    header_fill = PatternFill("solid", fgColor="1F4E79")
    header_font = Font(bold=True, color="FFFFFF")
    for col, title in enumerate(HEADERS, start=1):
        cell = ws.cell(row=1, column=col, value=title)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(wrap_text=True, vertical="center")


def write_findings(ws):
    rows = enrich_rows()
    style_header(ws)
    ws.title = "QA Report"
    for row_idx, row in enumerate(rows, start=2):
        for col_idx, value in enumerate(row, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.alignment = Alignment(wrap_text=True, vertical="top")
            if col_idx == 3 and value in SEVERITY_COLORS:
                cell.fill = PatternFill("solid", fgColor=SEVERITY_COLORS[value])
            if col_idx == 12 and value in STATUS_COLORS:
                cell.fill = PatternFill("solid", fgColor=STATUS_COLORS[value])
    widths = [10, 14, 10, 18, 28, 38, 28, 22, 22, 28, 34, 10, 42, 10]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = w
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{ws.cell(row=1, column=len(HEADERS)).column_letter}{len(rows)+1}"
    return rows


def write_summary(wb, rows: list[list[str]]):
    ws = wb.create_sheet("Summary", 0)
    ws["A1"] = "Billkaro Windows — QA Report (Updated After Fixes)"
    ws["A1"].font = Font(bold=True, size=16)

    status_counts: dict[str, int] = {}
    bug_status: dict[str, int] = {}
    for row in rows:
        st = row[11]
        status_counts[st] = status_counts.get(st, 0) + 1
        if row[1] == "Bug":
            bug_status[st] = bug_status.get(st, 0) + 1

    fixed_bugs = bug_status.get("Fixed", 0)
    open_bugs = bug_status.get("Open", 0) + bug_status.get("Deferred", 0)
    partial_bugs = bug_status.get("Partial", 0)
    total_bugs = sum(bug_status.values())

    info = [
        ("Report Date", datetime.now().strftime("%Y-%m-%d %H:%M")),
        ("Report Version", "v2 — Post-fix update"),
        ("Project", "billkaro_windows + billkaro-backend"),
        ("Platform", "Windows Desktop (Flutter 3.35.2)"),
        ("Build Status", "PASS — Release build succeeded"),
        ("Tests (billkaro_windows)", "1/1 pass (smoke test)"),
        ("Tests (billkaro-app)", "21/25 pass (4 legacy widget tests still failing)"),
        ("Total Findings", str(len(rows))),
        ("Bugs Fixed", f"{fixed_bugs} of {total_bugs}"),
        ("Bugs Partially Fixed", str(partial_bugs)),
        ("Bugs Still Open", str(open_bugs)),
        ("Backend Changes", "Staff login returns permissions; edit_menu in staff DTOs"),
    ]
    for i, (k, v) in enumerate(info, start=3):
        ws.cell(row=i, column=1, value=k).font = Font(bold=True)
        ws.cell(row=i, column=2, value=v)

    r = len(info) + 5
    ws.cell(row=r, column=1, value="Status Summary").font = Font(bold=True)
    r += 1
    for name in ["Fixed", "Partial", "Open", "Deferred"]:
        ws.cell(row=r, column=1, value=name)
        ws.cell(row=r, column=2, value=status_counts.get(name, 0))
        r += 1

    r += 1
    ws.cell(row=r, column=1, value="Remaining Open Bugs (verify manually)").font = Font(bold=True)
    r += 1
    for row in rows:
        if row[1] == "Bug" and row[11] in ("Open", "Partial"):
            ws.cell(row=r, column=1, value=f"{row[0]} — {row[4]}")
            ws.cell(row=r, column=2, value=row[11])
            r += 1

    r += 1
    ws.cell(row=r, column=1, value="Deployment Checklist").font = Font(bold=True)
    r += 1
    checklist = [
        "Set API_BASE_URL in .env to production URL",
        "Set Twilio credentials for WhatsApp marketing",
        "Restart backend to apply staff permissions in login response",
        "Re-save staff with edit_menu permission where needed",
        "Run flutter build windows and smoke-test login → order → sync",
    ]
    for item in checklist:
        ws.cell(row=r, column=1, value=f"• {item}")
        r += 1

    ws.column_dimensions["A"].width = 58
    ws.column_dimensions["B"].width = 62


def write_fixed_sheet(wb, rows: list[list[str]]):
    ws = wb.create_sheet("Fixed Items")
    style_header(ws)
    fixed = [r for r in rows if r[11] == "Fixed"]
    for row_idx, row in enumerate(fixed, start=2):
        for col_idx, value in enumerate(row, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.alignment = Alignment(wrap_text=True, vertical="top")
            if col_idx == 12:
                cell.fill = PatternFill("solid", fgColor=STATUS_COLORS["Fixed"])
    ws.freeze_panes = "A2"


def main():
    root = Path(__file__).resolve().parents[1]
    out_dir = root / "QA_Reports"
    out_dir.mkdir(exist_ok=True)
    stamp = datetime.now().strftime("%Y-%m-%d_%H%M")
    out_path = out_dir / f"Billkaro_QA_Report_Updated_{stamp}.xlsx"

    wb = Workbook()
    rows = write_findings(wb.active)
    write_summary(wb, rows)
    write_fixed_sheet(wb, rows)
    wb.save(out_path)
    print(f"Report saved: {out_path}")


if __name__ == "__main__":
    main()
